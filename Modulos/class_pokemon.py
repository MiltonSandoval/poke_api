import requests
import urllib3
from openpyxl import Workbook, load_workbook

#Esta funcion evita que se espanee el mensaje del ssl al hacer una consulta con el verify desactivado
urllib3.disable_warnings()

wb = load_workbook("pokemones.xlsx")
ws = wb.active


#Clase que recibe de parametro el nombre del pokemon y que tiene un metodo para buscar al pokemon en la api
class Pokemon:
    def __init__(self, nombre) -> None:
        self.nombre = nombre

    def Buscar(self):
        poke_dict = {}
        url = f"http://pokeapi.co/api/v2/pokemon/{self.nombre}/"
        req = requests.get(url, verify=False)
        Pokedex = req.json()
        poke_dict['name'] = Pokedex['name']
        poke_dict['id'] = Pokedex['id']
        poke_dict['tipo'] = Tipo(Pokedex['types'])
        return poke_dict

    def BuscarLocal(self):
        existe_pokemon = False
            # Itera solo sobre las filas con valores
        for fila in ws.iter_rows(values_only=True):
            for celda in fila:
                if str(celda) == self.nombre:
                    imprimir_poke(fila[0],fila[1],fila[2] )
                    existe_pokemon = True
                    break
        if not existe_pokemon:
            print(f"{self.nombre} no se encuentra en el Excel. Agregándolo...")
            poke_dict = self.Buscar()
            ws.append([poke_dict['id'], poke_dict['name'], poke_dict['tipo']])
            wb.save("pokemones.xlsx")
            self.BuscarLocal()

#Funcion que sirve para imprimir los tipos de los pokemones, en caso de que tenga mas de 1
def Tipo(tipos):
    tipo = []
    for elementos in tipos:
            tipo.append(elementos['type']['name'])
    tipo = ", ".join(tipo)
    return tipo


def imprimir_poke(name, id, tipo):
    print("-" * 20)
    print(f"Numero en la Pokedex: {id} \n"
        f"Nombre: {name} \n"
        f"Tipo: {tipo}")
    print("-" * 20)
